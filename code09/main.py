from openpyxl import load_workbook as load

DIR = 'D:\작업\CODE\Python3\code09\main.xlsx'

wb = load(DIR)

# ws = wb.create_sheet('test')
ws = wb['test']
# ws['A1'] = "제목1"
# ws['A2'] = "제목2"
a1 = ws['A2'].value
a2 = ws['B2'].value
# wb.save(DIR)
wb.close()